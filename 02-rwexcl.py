#excel数据自动读写

import openpyxl
import os
import time

#定义写入文件路径，创建保存文件名
def fullname():
	excelPath = os.path.join(os.getcwd(), 'ExcelData')
	print("***fullname保存路径测试***")
	print("Excel数据将保存路径为:%s"%excelPath)
    # 定义文件名称
    #  invalid mode ('wb') or filename: 'Excel2017-09-21_20:15:57.xlsx'   这种方式明明文件，会提示保存失败，无效的文件名。
    # nameTime = time.strftime('%Y-%m-%d_%H:%M:%S')
	nameTime = time.strftime('%Y-%m-%d_%H-%M-%S')
	excelName = 'Excel' + nameTime + '.xlsx' 
	ExcelFullName= os.path.join(excelPath,excelName)
	print('***fullname生成保存文件名测试****')
	print("Excel的文件名为：%s"%ExcelFullName)
	return(excelPath,excelName)


#写数据
	#定义表头
def def_head():
	tableTitle=[]
	var= int(input('观察项目数为（整数）：'))
	for i in range(1,var+1):
		s=input('输入第'+str(i)+'个观察项目：')
		tableTitle.append(s)
#return将生成的tableTitle的值返回给函数def_head(),以后可以调用此函数返回的值。如果没有return返回，调用时值就为NONE
	return(tableTitle)	
#	print('---def_head测试---')
#	print('观察项目tableTitle为：')
#	print(tableTitle)
	


	#写入表头
wb = openpyxl.Workbook()
ws = wb.create_sheet(index=0, title="mergfile")
ws = wb.active
def xlhead(alist,ws):
#	wb = openpyxl.Workbook()
#	ws = wb.create_sheet(index=0, title="mergfile")
#	ws = wb.active
#	if tableTitle(row) < 1 or tableTitle(col) < 1:
#		raise ValueError("Row or column values must be at least 1")
#	else:
	print('***xlhead测试***')
#alist为列表，ws为打开活动的工作表
	for col in range(len(alist)):
		c=col+1
		m=ws.cell(row=1,column=c).value=alist[col]
		print(m)



#写入数据
#slist原始数据为于表头相对应的list格式数据,ws为打开活动的工作表
def fildata(slist,ws):
	for row in range(len(slist)):
		ws.append(slist[row])
	print('fildata执行完毕')


#保存到生成的excel文件,filename为文件名
#指定保存路径格式：wb.save("F:/my_file/Python_files/others/{}".format(Marvel1))
def savexl(savepath,filename):
	if os.path.exists(savepath)==False:	
		os.makedirs(savepath)   #创建目录
	else:
		pass
	fulpath= savepath+"/{}"        
	wb.save(fulpath.format(filename))
	print('文件已保存在:',savepath)





if __name__=='__main__':
	pathandname=fullname()
	savepath=pathandname[0]
	filename=pathandname[1]
	print(savepath,filename)
#	tableTitle = ['userName', 'Phone', 'age', 'Remark']
	tableTitle=def_head()
	print(tableTitle)
	xlhead(tableTitle,ws)
	slist=[['张学友', 15201062100, 18, '测试数据！'], ['李雷', 15201062598, 19, '测试数据！'],['Marry', 15201062191, 28, '测试数据！']]
	fildata(slist,ws)
	savexl(savepath,filename)











#读文件


#if __name__ == '__main__':
#	fullname()
#    ExcelFullName = writeExcel()
#    readExcel(ExcelFullName)






